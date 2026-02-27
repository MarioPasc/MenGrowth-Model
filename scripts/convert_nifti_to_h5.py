#!/usr/bin/env python
# scripts/convert_nifti_to_h5.py
"""Convert BraTS-MEN NIfTI files to a single HDF5 file.

Pre-applies spatial preprocessing (Orient→Resample→CropForeground→SpatialPad→
CenterCrop at 192³) but does NOT apply z-score normalization. Normalization
happens at runtime so augmentation operates on the original intensity scale.

H5 Schema:
    brats_men_train.h5
    ├── attrs: {n_subjects, roi_size, spacing, channel_order, version}
    ├── images          [N, 4, 192, 192, 192] float32  chunks=[1,...] gzip
    ├── segs            [N, 1, 192, 192, 192] int8      chunks=[1,...] gzip
    ├── subject_ids     [N] string (sorted alphabetically)
    ├── metadata/
    │   ├── grade       [N] int8  (-1 for missing)
    │   ├── age         [N] float32 (NaN for missing)
    │   └── sex         [N] string
    ├── semantic/
    │   ├── volume      [N, 4] float32
    │   ├── location    [N, 3] float32
    │   └── shape       [N, 3] float32
    └── splits/
        ├── lora_train  [525] int32
        ├── lora_val    [100] int32
        ├── sdp_train   [225] int32
        └── test        [150] int32

Usage:
    # Dry-run with 5 subjects
    python scripts/convert_nifti_to_h5.py \
        --data-root /path/to/BraTS_Men_Train \
        --output brats_men_train.h5 \
        --max-subjects 5

    # Full conversion (all 1000 subjects)
    python scripts/convert_nifti_to_h5.py \
        --data-root /path/to/BraTS_Men_Train \
        --output brats_men_train.h5

Estimated output size: ~15 GB for 1000 subjects (gzip-compressed).
"""

import argparse
import logging
import sys
from pathlib import Path

import h5py
import numpy as np
from monai.transforms import (
    Compose,
    ConcatItemsd,
    CropForegroundd,
    EnsureChannelFirstd,
    EnsureTyped,
    LoadImaged,
    Orientationd,
    ResizeWithPadOrCropd,
    Spacingd,
    SpatialPadd,
)
from tqdm import tqdm

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from growth.data.bratsmendata import (
    BraTSMENDataset,
    split_subjects_multi,
)
from growth.data.semantic_features import extract_semantic_features, extract_semantic_features_from_file
from growth.data.transforms import (
    DEFAULT_SPACING,
    FEATURE_ROI_SIZE,
    MODALITY_KEYS,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# H5 file version
H5_VERSION = "1.0"

# Default split sizes (matching the pipeline's 4-way split)
DEFAULT_SPLIT_SIZES = {
    "lora_train": 525,
    "lora_val": 100,
    "sdp_train": 225,
    "test": 150,
}


def build_preprocessing_transforms() -> Compose:
    """Build MONAI transforms for spatial preprocessing without normalization.

    Pipeline: Load → EnsureChannelFirst → Orient(RAS) → Spacing(1mm) →
              CropForeground → SpatialPad(192³) → CenterCrop(192³) →
              Concat → EnsureType

    Returns:
        MONAI Compose pipeline.
    """
    modality_keys = list(MODALITY_KEYS)
    seg_key = "seg"
    all_keys = modality_keys + [seg_key]
    roi_size = list(FEATURE_ROI_SIZE)

    transforms = [
        # Load NIfTI files
        LoadImaged(keys=all_keys, image_only=False),
        EnsureChannelFirstd(keys=all_keys),
        # Reorient to RAS
        Orientationd(keys=all_keys, axcodes="RAS"),
        # Resample to 1mm isotropic
        Spacingd(
            keys=modality_keys,
            pixdim=DEFAULT_SPACING,
            mode=["bilinear"] * len(modality_keys),
        ),
        Spacingd(
            keys=[seg_key],
            pixdim=DEFAULT_SPACING,
            mode=("nearest",),
        ),
        # Crop to brain foreground
        CropForegroundd(
            keys=all_keys,
            source_key=modality_keys[0],
            k_divisible=roi_size,
        ),
        # Pad to at least 192³
        SpatialPadd(keys=all_keys, spatial_size=roi_size),
        # Center crop to exactly 192³
        ResizeWithPadOrCropd(keys=all_keys, spatial_size=roi_size),
        # Concatenate modalities: [t2f, t1c, t1n, t2w] → [4, 192, 192, 192]
        ConcatItemsd(keys=modality_keys, name="image", dim=0),
        # Convert to tensors (no normalization!)
        EnsureTyped(keys=["image", seg_key], dtype="float32", track_meta=False),
    ]

    return Compose(transforms)


def load_metadata(data_root: Path) -> dict[str, dict]:
    """Load clinical metadata from XLSX file if available.

    Args:
        data_root: BraTS-MEN data root directory.

    Returns:
        Dict mapping subject_id → {grade, age, sex}.
    """
    # Try to find metadata XLSX
    xlsx_candidates = list(data_root.glob("*.xlsx")) + list(
        data_root.parent.glob("*meningioma*.xlsx")
    )

    if not xlsx_candidates:
        logger.warning("No metadata XLSX found. Using defaults (-1/NaN/'unknown').")
        return {}

    try:
        import openpyxl

        xlsx_path = xlsx_candidates[0]
        logger.info(f"Loading metadata from {xlsx_path}")

        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        ws = wb.active

        # Read header row
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # Find relevant columns (case-insensitive)
        header_lower = [h.lower() if h else "" for h in header]
        id_col = None
        grade_col = None
        age_col = None
        sex_col = None

        for i, h in enumerate(header_lower):
            if "id" in h or "subject" in h or "brats" in h:
                id_col = i
            elif "grade" in h:
                grade_col = i
            elif "age" in h:
                age_col = i
            elif "sex" in h or "gender" in h:
                sex_col = i

        if id_col is None:
            logger.warning("Could not find ID column in XLSX. Skipping metadata.")
            return {}

        metadata = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            subject_id = str(row[id_col]) if row[id_col] else None
            if not subject_id:
                continue

            grade = (
                int(row[grade_col]) if grade_col is not None and row[grade_col] is not None else -1
            )
            age = (
                float(row[age_col])
                if age_col is not None and row[age_col] is not None
                else float("nan")
            )
            sex = (
                str(row[sex_col]) if sex_col is not None and row[sex_col] is not None else "unknown"
            )

            metadata[subject_id] = {"grade": grade, "age": age, "sex": sex}

        wb.close()
        logger.info(f"Loaded metadata for {len(metadata)} subjects")
        return metadata

    except ImportError:
        logger.warning("openpyxl not installed. Skipping metadata.")
        return {}
    except Exception as e:
        logger.warning(f"Failed to load metadata: {e}. Using defaults.")
        return {}


def convert(
    data_root: str,
    output_path: str,
    max_subjects: int | None = None,
    seed: int = 42,
    compression: str = "gzip",
    compression_level: int = 4,
) -> None:
    """Convert NIfTI files to HDF5.

    Args:
        data_root: Path to BraTS_Men_Train directory.
        output_path: Output H5 file path.
        max_subjects: Max subjects for dry-run (None = all).
        seed: Random seed for split generation.
        compression: HDF5 compression algorithm.
        compression_level: Compression level (1-9).
    """
    data_root = Path(data_root)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Discover subjects
    all_subjects = BraTSMENDataset.get_all_subject_ids(data_root)
    logger.info(f"Found {len(all_subjects)} subjects in {data_root}")

    if max_subjects is not None:
        all_subjects = all_subjects[:max_subjects]
        logger.info(f"Limiting to {max_subjects} subjects (dry-run)")

    n_subjects = len(all_subjects)
    roi = list(FEATURE_ROI_SIZE)

    # Load metadata
    metadata = load_metadata(data_root)

    # Build preprocessing pipeline (no normalization)
    transforms = build_preprocessing_transforms()

    # Generate splits
    if max_subjects is None or max_subjects >= sum(DEFAULT_SPLIT_SIZES.values()):
        split_sizes = DEFAULT_SPLIT_SIZES
    else:
        # For dry-run: proportional splits
        total_default = sum(DEFAULT_SPLIT_SIZES.values())
        split_sizes = {}
        allocated = 0
        for name, size in DEFAULT_SPLIT_SIZES.items():
            if name == list(DEFAULT_SPLIT_SIZES.keys())[-1]:
                split_sizes[name] = n_subjects - allocated
            else:
                split_sizes[name] = max(1, round(size * n_subjects / total_default))
                allocated += split_sizes[name]

    splits = split_subjects_multi(all_subjects, split_sizes, seed=seed)

    # Build subject → index mapping (subjects are stored in alphabetical order)
    subject_to_idx = {sid: i for i, sid in enumerate(all_subjects)}

    # Create H5 file
    logger.info(f"Creating H5 file: {output_path}")
    logger.info(f"  Subjects: {n_subjects}")
    logger.info(f"  ROI size: {roi}")
    logger.info(f"  Compression: {compression} level={compression_level}")

    with h5py.File(output_path, "w") as f:
        # Global attributes
        f.attrs["n_subjects"] = n_subjects
        f.attrs["roi_size"] = roi
        f.attrs["spacing"] = list(DEFAULT_SPACING)
        f.attrs["channel_order"] = list(MODALITY_KEYS)
        f.attrs["version"] = H5_VERSION

        # Pre-allocate datasets
        images_ds = f.create_dataset(
            "images",
            shape=(n_subjects, 4, *roi),
            dtype="float32",
            chunks=(1, 4, *roi),
            compression=compression,
            compression_opts=compression_level,
        )

        segs_ds = f.create_dataset(
            "segs",
            shape=(n_subjects, 1, *roi),
            dtype="int8",
            chunks=(1, 1, *roi),
            compression=compression,
            compression_opts=compression_level,
        )

        # Subject IDs (variable-length strings)
        dt = h5py.special_dtype(vlen=str)
        f.create_dataset("subject_ids", data=all_subjects, dtype=dt)

        # Metadata group
        meta_grp = f.create_group("metadata")
        grade_ds = meta_grp.create_dataset("grade", shape=(n_subjects,), dtype="int8")
        age_ds = meta_grp.create_dataset("age", shape=(n_subjects,), dtype="float32")
        sex_ds = meta_grp.create_dataset("sex", shape=(n_subjects,), dtype=dt)

        # Semantic features group
        sem_grp = f.create_group("semantic")
        vol_ds = sem_grp.create_dataset("volume", shape=(n_subjects, 4), dtype="float32")
        loc_ds = sem_grp.create_dataset("location", shape=(n_subjects, 3), dtype="float32")
        shape_ds = sem_grp.create_dataset("shape", shape=(n_subjects, 3), dtype="float32")

        # Process each subject
        n_errors = 0
        for i, subject_id in enumerate(tqdm(all_subjects, desc="Converting")):
            try:
                # Get file paths
                paths = BraTSMENDataset.load_subject_paths(data_root, subject_id)

                # Apply spatial preprocessing (no normalization)
                data = {
                    "t2f": str(paths["t2f"]),
                    "t1c": str(paths["t1c"]),
                    "t1n": str(paths["t1n"]),
                    "t2w": str(paths["t2w"]),
                    "seg": str(paths["seg"]),
                }
                result = transforms(data)

                # Write image and seg
                image_np = result["image"].numpy()  # [4, 192, 192, 192]
                seg_np = result["seg"].numpy()  # [1, 192, 192, 192]

                assert image_np.shape == (4, *roi), f"Image shape mismatch: {image_np.shape}"
                assert seg_np.shape == (1, *roi), f"Seg shape mismatch: {seg_np.shape}"

                images_ds[i] = image_np
                segs_ds[i] = seg_np.astype(np.int8)

                # Metadata
                meta = metadata.get(subject_id, {})
                grade_ds[i] = meta.get("grade", -1)
                age_ds[i] = meta.get("age", float("nan"))
                sex_ds[i] = meta.get("sex", "unknown")

                # Semantic features from preprocessed 192³ seg (consistent frame)
                seg_for_semantic = seg_np[0].astype(np.int32)  # [192, 192, 192]
                sem = extract_semantic_features(seg_for_semantic, spacing=(1.0, 1.0, 1.0))
                vol_ds[i] = sem["volume"]
                loc_ds[i] = sem["location"]
                shape_ds[i] = sem["shape"]

            except Exception as e:
                logger.error(f"Failed to process {subject_id}: {e}")
                n_errors += 1
                continue

        # Splits group (store as index arrays)
        splits_grp = f.create_group("splits")
        for split_name, subject_list in splits.items():
            indices = np.array([subject_to_idx[sid] for sid in subject_list], dtype=np.int32)
            splits_grp.create_dataset(split_name, data=indices)
            logger.info(f"  Split '{split_name}': {len(indices)} subjects")

    # Summary
    file_size_gb = output_path.stat().st_size / (1024**3)
    logger.info("\nConversion complete!")
    logger.info(f"  Output: {output_path}")
    logger.info(f"  File size: {file_size_gb:.2f} GB")
    logger.info(f"  Subjects: {n_subjects} ({n_errors} errors)")

    # Verification
    _verify_h5(output_path, all_subjects, n_subjects, roi)


def _verify_h5(
    h5_path: Path,
    subject_ids: list[str],
    n_subjects: int,
    roi: list[int],
) -> None:
    """Spot-check the H5 file for correctness."""
    logger.info("\nVerification:")

    with h5py.File(h5_path, "r") as f:
        # Check attributes
        assert f.attrs["n_subjects"] == n_subjects
        assert list(f.attrs["roi_size"]) == roi
        assert list(f.attrs["channel_order"]) == list(MODALITY_KEYS)

        # Check shapes
        assert f["images"].shape == (n_subjects, 4, *roi)
        assert f["segs"].shape == (n_subjects, 1, *roi)
        assert len(f["subject_ids"]) == n_subjects

        # Check semantic shapes
        assert f["semantic/volume"].shape == (n_subjects, 4)
        assert f["semantic/location"].shape == (n_subjects, 3)
        assert f["semantic/shape"].shape == (n_subjects, 3)

        # Spot-check 3 random subjects
        rng = np.random.RandomState(42)
        check_indices = rng.choice(n_subjects, min(3, n_subjects), replace=False)

        for idx in check_indices:
            sid = f["subject_ids"][idx]
            if isinstance(sid, bytes):
                sid = sid.decode()

            img = f["images"][idx]
            seg = f["segs"][idx]

            assert not np.isnan(img).any(), f"NaN in image for {sid}"
            assert not np.isinf(img).any(), f"Inf in image for {sid}"
            assert img.max() > 0, f"All-zero image for {sid}"

            unique_labels = np.unique(seg)
            valid_labels = {0, 1, 2, 3}
            assert set(unique_labels).issubset(valid_labels), (
                f"Invalid labels {unique_labels} for {sid}"
            )

            logger.info(
                f"  [{idx}] {sid}: "
                f"img range=[{img.min():.1f}, {img.max():.1f}], "
                f"seg labels={sorted(unique_labels)}"
            )

        # Check splits sum
        if "splits" in f:
            total_in_splits = sum(len(f[f"splits/{s}"]) for s in f["splits"])
            logger.info(f"  Total subjects in splits: {total_in_splits}")

    logger.info("  Verification passed!")


def main() -> None:
    """Main entry point."""
    parser = argparse.ArgumentParser(description="Convert BraTS-MEN NIfTI files to HDF5")
    parser.add_argument(
        "--data-root",
        type=str,
        required=True,
        help="Path to BraTS_Men_Train directory",
    )
    parser.add_argument(
        "--output",
        type=str,
        default="brats_men_train.h5",
        help="Output H5 file path",
    )
    parser.add_argument(
        "--max-subjects",
        type=int,
        default=None,
        help="Max subjects for dry-run testing (default: all)",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="Random seed for split generation",
    )
    parser.add_argument(
        "--compression-level",
        type=int,
        default=4,
        help="Gzip compression level 1-9 (default: 4)",
    )

    args = parser.parse_args()

    convert(
        data_root=args.data_root,
        output_path=args.output,
        max_subjects=args.max_subjects,
        seed=args.seed,
        compression_level=args.compression_level,
    )


if __name__ == "__main__":
    main()
